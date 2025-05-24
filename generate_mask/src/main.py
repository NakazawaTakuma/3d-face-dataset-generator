#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import argparse
import random

import cv2
import cupy as cp
import openpyxl as excel
import pandas as pd


from modules.set_age_gender import set_age_gender
from modules.set_skin_rgb import set_skin_rgb
from modules.set_lip_rgb import set_lip_rgb
from modules.set_redness import set_redness
from modules.set_mark import set_mark
from modules.set_freckles import set_freckles
from modules.set_wrinkles_01_brow import set_wrinkles_01_brow
from modules.set_wrinkles_02_eyebrows1 import set_wrinkles_02_eyebrows1
from modules.set_wrinkles_03_glabella import set_wrinkles_03_glabella
from modules.set_wrinkles_04_double_eye import set_wrinkles_04_double_eye
from modules.set_wrinkles_05_under_eye import set_wrinkles_05_under_eye
from modules.set_wrinkles_06_nasolabial import set_wrinkles_06_nasolabial
from modules.set_wrinkles_07_mouth import set_wrinkles_07_mouth
from modules.set_wrinkles_08_neck import set_wrinkles_08_neck
from modules.set_cosmetic import set_cosmetic
from modules.set_hair_and_eye_rgb import set_hair_and_eye_rgb
from modules.set_hair_brow import set_hair_brow


# GPU メモリプールの初期化（必要に応じてコメントアウト可）
cp.cuda.set_allocator(cp.cuda.MemoryPool().malloc)


def make_dir_if_not_exists(path: str):
    os.makedirs(path, exist_ok=True)


def remove_file_if_exists(path: str):
    if os.path.exists(path):
        os.remove(path)


def write_2d_list_to_sheet(ws, data_2d, start_row: int, start_col: int):
    """2次元リストを Excel ワークシートに書き込むヘルパー関数"""
    for r, row in enumerate(data_2d):
        for c, val in enumerate(row):
            ws.cell(row=start_row + r, column=start_col + c, value=val)


def load_mask_images(asset_dir: str, img_size: int):
    """
    asset_dir/mask フォルダにある各種マスク画像を読み込み、
    前処理（リサイズ＋正規化）してリストに詰める
    """
    def load_gray(path: str, size: int):
        img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
        if img is None:
            raise FileNotFoundError(f"ファイルが見つかりません: {path}")
        img = cv2.resize(img, (size, size))
        return (cp.array(img) / 255.0).astype(cp.float32)

    base = os.path.join(asset_dir, "mask")
    load_list = [
        "remove.png",
        "wrinkles_01_brow_remove.png",
        "wrinkles_01_brow.png",
        "wrinkles_02_eyebrows00.png",
        "wrinkles_02_eyebrows06.png",
        "wrinkles_02_E_eyebrows00.png",
        "wrinkles_02_E_eyebrows01.png",
        "wrinkles_03_glabella00.png",
        "wrinkles_04_double_eye00.png",
        "wrinkles_04_double_eye01.png",
        "wrinkles_04_under_eye_remove.png",
        "wrinkles_04_under_eye00.png",
        "wrinkles_04_under_eye01.png",
        "wrinkles_04_under_eye03.png",
        "wrinkles_05_nasolabial02.png",
        "wrinkles_05_nasolabial03.png",
        "wrinkles_05_nasolabial04.png",
        "wrinkles_06_mouth00.png",
        "wrinkles_07_neck00.png",
        "cosmetic_01_eyeshadow_above00.png",
        "cosmetic_01_eyeshadow_above01.png",
        "cosmetic_01_eyeshadow_below00.png",
        "cosmetic_01_eyeshadow_below01.png",
        "cosmetic_remove.png",
        "cosmetic_02_eyeline00.png",
        "hair_base.png",
        "hair_brow.png",
    ]
    images = []
    for filename in load_list:
        fullpath = os.path.join(base, filename)
        images.append(load_gray(fullpath, img_size))
    return images


def create_uv_data(lp: int, c_data, c_img, img_size: int, use_image):
    """１サンプル分の UV マスクデータと属性を作成"""
    # 年齢・性別
    set_age_gender(lp, c_data)

    # 太り具合（幼児は確率的にスキップ）
    far_rate_02 = 90
    far_rate_03 = 96
    infant_age = 3
    if c_data[lp][2] < infant_age and random.randint(0, 100) > 0:
        c_data[lp][20] = random.randint(far_rate_02, far_rate_03 - 1)
    else:
        c_data[lp][20] = random.randint(1, 100)

    # 肌・唇・赤み
    set_skin_rgb(lp, c_data)
    set_lip_rgb(lp, c_data)
    set_redness(lp, c_data)

    # マーク（ほくろ、そばかすなど）
    set_mark(lp, c_data, c_img, img_size, use_image)
    set_freckles(lp, c_img, img_size, use_image)

    # しわ
    set_wrinkles_01_brow(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_02_eyebrows1(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_03_glabella(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_04_double_eye(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_05_under_eye(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_06_nasolabial(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_07_mouth(lp, c_data, c_img, img_size, use_image)
    set_wrinkles_08_neck(lp, c_data, c_img, img_size, use_image)

    # 化粧
    set_cosmetic(lp, c_data, c_img, img_size, use_image)

    # 髪・目
    set_hair_and_eye_rgb(lp, c_data)
    set_hair_brow(lp, c_img, img_size, use_image)

    print(f"  サンプル {lp} 処理完了")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.abspath(os.path.join(script_dir, os.pardir))
    # ───────────────────────────────────────────────────
    # ↓ デフォルト値をここで定義しておけば、引数を指定しなくてもこれが使われる
    DEFAULT_ASSET_DIR = parent_dir + "/assets"
    DEFAULT_OUTPUT_DIR = parent_dir + "/output/mask"
    DEFAULT_NUM_SAMPLES = 50
    DEFAULT_BATCH_SIZE = 10
    DEFAULT_IMG_SIZE = 2048
    # ↑───────────────────────────────────────────────────

    parser = argparse.ArgumentParser(
        description="GenerateFaceModel: 顔マスク画像 + 属性データ を生成します"
    )
    parser.add_argument(
        "--asset-dir",
        type=str,
        default=DEFAULT_ASSET_DIR,
        help=f"マスク画像テンプレートのディレクトリ（デフォルト: {DEFAULT_ASSET_DIR}）",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=DEFAULT_OUTPUT_DIR,
        help=f"出力ディレクトリ（デフォルト: {DEFAULT_OUTPUT_DIR}）",
    )
    parser.add_argument(
        "--num-samples", "-n",
        type=int,
        default=DEFAULT_NUM_SAMPLES,
        help=f"生成サンプル数（デフォルト: {DEFAULT_NUM_SAMPLES}）",
    )
    parser.add_argument(
        "--batch-size", "-b",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help=f"バッチサイズ（デフォルト: {DEFAULT_BATCH_SIZE}）",
    )
    parser.add_argument(
        "--img-size",
        type=int,
        default=DEFAULT_IMG_SIZE,
        help=f"マスク画像サイズ（正方形、デフォルト: {DEFAULT_IMG_SIZE}）",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="既存の dataset.xlsx と dataset.csv を上書きする場合に指定",
    )
    args = parser.parse_args()

    # 引数チェック
    total = args.num_samples
    batch = args.batch_size
    if total % batch != 0:
        parser.error("`--num-samples` は `--batch-size` の倍数にしてください。")

    img_size = args.img_size


    # 出力先フォルダの用意
    excel_name = "dataset.xlsx"
    csv_name = "dataset.csv"
    make_dir_if_not_exists(args.output_dir)

    excel_path = os.path.join(args.output_dir, excel_name)
    csv_path = os.path.join(args.output_dir, csv_name)

    if args.overwrite:
        remove_file_if_exists(excel_path)
        remove_file_if_exists(csv_path)


    categories = [
        "lp", "gender", "age",
        "skin_color_R", "skin_color_G", "skin_color_B",
        "lips_color_R", "lips_color_G", "lips_color_B",
        "use_lip", "skin_red_cheek", "skin_red_cheek2",
        "skin_red_jaw", "skin_red_forehead", "skin_red_nose",
        "skin_red_center", "skin_red_ear", "double_eye",
        "skin_light", "nasolabial", "fat_rate", "makeup",
        "eyeshadow_color_R", "eyeshadow_color_G", "eyeshadow_color_B",
        "hair_color1_R", "hair_color1_G", "hair_color1_B",
        "hair_color2_R", "hair_color2_G", "hair_color2_B",
        "hair_bread_color1_R", "hair_bread_color1_G", "hair_bread_color1_B",
        "hair_bread_color2_R", "hair_bread_color2_G", "hair_bread_color2_B",
        "hair_brows_color_R", "hair_brows_color_G", "hair_brows_color_B",
        "inline_eye", "skin_red_eyeup", "skin_red_eyedown",
        "eyelashes_length_up", "eyelashes_length_down", "eyelashes_angle", "eyelashes_density",
        "eyelashes_R", "eyelashes_G", "eyelashes_B",
        "eye_R_1_R", "eye_R_1_G", "eye_R_1_B",
        "eye_R_2_R", "eye_R_2_G", "eye_R_2_B",
        "eye_L_1_R", "eye_L_1_G", "eye_L_1_B",
        "eye_L_2_R", "eye_L_2_G", "eye_L_2_B",
        "eye_size", "eye_black_size"
    ]


    # Excel ワークブックの準備
    if not os.path.exists(excel_path):
        wb = excel.Workbook()
        ws = wb.active


        # ヘッダ行書き込み
        for idx, cat in enumerate(categories, start=1):
            ws.cell(row=1, column=idx, value=cat)
        start_step = 0
    else:
        wb = excel.load_workbook(excel_path)
        ws = wb.active
        # 既存の行数から開始行を決定（ヘッダ行を除いて、2行目以降がデータ）
        start_step = ws.max_row - 1

    # マスク画像格納フォルダの準備
    mask_categories = {
        "mark": ["mark", "mark_black", "acne", "mole", "wart", "freckles"],
        "wrinkles": [
            "01_brow", "01_brow_E",
            "02_eyebrows1", "02_eyebrows1_E",
            "03_glabella", "03_glabella_E",
            "04_double_eye", "04_double_eye2",
            "05_under_eye", "05_under_eye2",
            "06_nasolabial_v", "06_nasolabial_^",
            "06_nasolabial_v_E", "06_nasolabial_^_E",
            "07_mouth", "08_neck"
        ],
        "cosmetic": ["01_eyeshadow", "02_eyeline"],
        "hair": ["01_hair", "02_bread", "03_brow"]
    }
    img_category_num = sum(len(v) for v in mask_categories.values())

    for cat_main, sub_list in mask_categories.items():
        for sub in sub_list:
            output_subdir = os.path.join(args.output_dir, cat_main, sub)
            make_dir_if_not_exists(output_subdir)

    # マスク画像の読み込み
    use_image = load_mask_images(args.asset_dir, img_size)

    # メインループ（バッチごとに処理）
    for batch_idx in range(total // batch):
        print(f"[Batch {batch_idx + 1}/{total // batch}] 生成中…")

        # バッチ用のデータ配列と画像配列を作成
        c_data = cp.empty((batch, len(categories)), dtype=cp.float32)
        c_img = cp.zeros((batch, img_category_num, img_size, img_size), dtype=cp.float32)

        # 各サンプルを生成
        for i in range(batch):
            global_lp = i + batch * batch_idx
            c_data[i][0] = global_lp  # lp 番号
            create_uv_data(i, c_data, c_img, img_size, use_image)

        # 画像を保存
        for i in range(batch):
            idx_counter = 0
            for cat_main, sub_list in mask_categories.items():
                for sub in sub_list:
                    img_np = cp.asnumpy(c_img[i][idx_counter])
                    img_resized = cv2.resize(img_np, (img_size * 2, img_size * 2))
                    filename = f"{i + batch * batch_idx}.png"
                    out_path = os.path.join(args.output_dir, cat_main, sub, filename)
                    cv2.imwrite(out_path, img_resized)
                    idx_counter += 1

        # データを Excel に追記
        write_2d_list_to_sheet(ws,
                              cp.asnumpy(c_data).tolist(),
                              start_row=start_step + 2 + batch * batch_idx,
                              start_col=1)

        wb.save(excel_path)

        # Excel を読み直して CSV に変換（追記）
        df = pd.read_excel(excel_path)
        df.to_csv(csv_path, index=False)

    print("全バッチ完了。出力先:", args.output_dir)


if __name__ == "__main__":
    main()
